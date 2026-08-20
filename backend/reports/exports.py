import json
from datetime import datetime
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.http import FileResponse, HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
    SimpleDocTemplate,
)
from reportlab.graphics.shapes import Drawing, String, Rect
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie

from .utils import build_report, get_date_range


# ============================================================
# BUDGETBUDDY PROFESSIONAL PDF THEME
# ============================================================

NAVY = HexColor("#0F172A")
NAVY_2 = HexColor("#172033")
SLATE = HexColor("#475569")
MUTED = HexColor("#64748B")
BORDER = HexColor("#CBD5E1")
LIGHT = HexColor("#F8FAFC")
LIGHT_BLUE = HexColor("#EFF6FF")

GREEN = HexColor("#16A34A")
BLUE = HexColor("#2563EB")
RED = HexColor("#DC2626")
PURPLE = HexColor("#9333EA")
AMBER = HexColor("#D97706")
CYAN = HexColor("#0891B2")


# ============================================================
# FONT SUPPORT
# Prevents the black-square problem that appeared before ₹.
# ============================================================

FONT_REGULAR = "Helvetica"
FONT_BOLD = "Helvetica-Bold"
CURRENCY_PREFIX = "Rs. "

font_candidates = [
    Path(r"C:\Windows\Fonts\DejaVuSans.ttf"),
    Path(r"C:\Windows\Fonts\DejaVuSans-Bold.ttf"),
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
]

regular_font = next(
    (p for p in font_candidates if p.name == "DejaVuSans.ttf" and p.exists()),
    None,
)
bold_font = next(
    (p for p in font_candidates if p.name == "DejaVuSans-Bold.ttf" and p.exists()),
    None,
)

if regular_font and bold_font:
    try:
        pdfmetrics.registerFont(
            TTFont("BudgetBuddySans", str(regular_font))
        )
        pdfmetrics.registerFont(
            TTFont("BudgetBuddySansBold", str(bold_font))
        )
        FONT_REGULAR = "BudgetBuddySans"
        FONT_BOLD = "BudgetBuddySansBold"
        CURRENCY_PREFIX = "₹"
    except Exception:
        pass


# ============================================================
# BRAND ASSET
# ============================================================

def get_budgetbuddy_logo_path():
    """Return the shared BudgetBuddy mark if it exists."""
    base_dir = Path(
        getattr(
            settings,
            "BASE_DIR",
            Path(__file__).resolve().parents[2],
        )
    )

    candidates = [
        base_dir / "frontend" / "public" / "budgetbuddy-mark.png",
        base_dir.parent / "frontend" / "public" / "budgetbuddy-mark.png",
        base_dir / "static" / "budgetbuddy-mark.png",
        base_dir / "static" / "images" / "budgetbuddy-mark.png",
        base_dir / "staticfiles" / "budgetbuddy-mark.png",
        Path(__file__).resolve().parent / "assets" / "budgetbuddy-mark.png",
    ]

    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate

    return None


# ============================================================
# HELPERS
# ============================================================

def number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def money(value):
    return f"{CURRENCY_PREFIX}{number(value):,.2f}"


def percent(value):
    return f"{number(value):.1f}%"


def safe_text(value, fallback="—"):
    if value is None or str(value).strip() == "":
        return fallback
    return str(value)


def styles():
    base = getSampleStyleSheet()

    return {
        "title": ParagraphStyle(
            "BBTitle",
            parent=base["Title"],
            fontName=FONT_BOLD,
            fontSize=23,
            leading=26,
            textColor=colors.white,
            alignment=TA_LEFT,
        ),
        "subtitle": ParagraphStyle(
            "BBSubtitle",
            parent=base["Normal"],
            fontName=FONT_REGULAR,
            fontSize=8.5,
            leading=11,
            textColor=HexColor("#CBD5E1"),
        ),
        "section": ParagraphStyle(
            "BBSection",
            parent=base["Heading2"],
            fontName=FONT_BOLD,
            fontSize=13,
            leading=16,
            textColor=NAVY,
            spaceAfter=3,
        ),
        "body": ParagraphStyle(
            "BBBody",
            parent=base["BodyText"],
            fontName=FONT_REGULAR,
            fontSize=8.5,
            leading=12,
            textColor=NAVY,
        ),
        "body_bold": ParagraphStyle(
            "BBBodyBold",
            parent=base["BodyText"],
            fontName=FONT_BOLD,
            fontSize=8.5,
            leading=12,
            textColor=NAVY,
        ),
        "muted": ParagraphStyle(
            "BBMuted",
            parent=base["BodyText"],
            fontName=FONT_REGULAR,
            fontSize=7.3,
            leading=10,
            textColor=MUTED,
        ),
        "card_label": ParagraphStyle(
            "BBCardLabel",
            parent=base["BodyText"],
            fontName=FONT_REGULAR,
            fontSize=7.1,
            leading=9,
            textColor=MUTED,
        ),
        "card_value": ParagraphStyle(
            "BBCardValue",
            parent=base["BodyText"],
            fontName=FONT_BOLD,
            fontSize=13.5,
            leading=16,
            textColor=NAVY,
        ),
        "white_small": ParagraphStyle(
            "BBWhiteSmall",
            parent=base["BodyText"],
            fontName=FONT_REGULAR,
            fontSize=7.5,
            leading=10,
            textColor=HexColor("#CBD5E1"),
        ),
        "white_bold": ParagraphStyle(
            "BBWhiteBold",
            parent=base["BodyText"],
            fontName=FONT_BOLD,
            fontSize=9,
            leading=11,
            textColor=colors.white,
        ),
        "center": ParagraphStyle(
            "BBCenter",
            parent=base["BodyText"],
            fontName=FONT_REGULAR,
            fontSize=7.5,
            leading=10,
            textColor=MUTED,
            alignment=TA_CENTER,
        ),
    }


def section_header(text, accent=BLUE):
    s = styles()
    t = Table(
        [[Paragraph(text, s["section"])]],
        colWidths=[174 * mm],
    )
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
        ("LINEBEFORE", (0, 0), (0, 0), 3, accent),
        ("BOX", (0, 0), (-1, -1), 0.4, BORDER),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


def stat_card(label, value, accent):
    s = styles()
    t = Table(
        [
            [Paragraph(label.upper(), s["card_label"])],
            [Paragraph(value, s["card_value"])],
        ],
        colWidths=[41.5 * mm],
    )
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("LINEABOVE", (0, 0), (-1, 0), 3, accent),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    return t


# ============================================================
# CHARTS
# ============================================================

def income_expense_chart(income, expense):
    drawing = Drawing(500, 170)

    chart = VerticalBarChart()
    chart.x = 60
    chart.y = 32
    chart.width = 410
    chart.height = 110

    chart.data = [[income, expense]]
    chart.categoryAxis.categoryNames = ["Income", "Expenses"]

    max_value = max(income, expense, 1)
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = max_value * 1.2
    chart.valueAxis.valueStep = max(max_value / 4, 1)

    chart.valueAxis.labels.fontName = FONT_REGULAR
    chart.valueAxis.labels.fontSize = 7
    chart.valueAxis.labels.fillColor = MUTED

    chart.categoryAxis.labels.fontName = FONT_REGULAR
    chart.categoryAxis.labels.fontSize = 8
    chart.categoryAxis.labels.fillColor = NAVY

    chart.valueAxis.visibleGrid = True
    chart.valueAxis.gridStrokeColor = HexColor("#E2E8F0")
    chart.valueAxis.gridStrokeWidth = 0.5

    chart.bars[0].fillColor = BLUE
    chart.bars[0].strokeColor = BLUE

    drawing.add(chart)

    drawing.add(
        String(
            250,
            155,
            "Income vs Expenses",
            textAnchor="middle",
            fontName=FONT_BOLD,
            fontSize=10,
            fillColor=NAVY,
        )
    )

    return drawing


def expense_donut(category_rows):
    drawing = Drawing(500, 185)

    rows = [
        r for r in (category_rows or [])
        if number(r.get("total", 0)) > 0
    ]

    if not rows:
        drawing.add(
            String(
                250,
                90,
                "No expense data available",
                textAnchor="middle",
                fontName=FONT_REGULAR,
                fontSize=9,
                fillColor=MUTED,
            )
        )
        return drawing

    pie = Pie()
    pie.x = 35
    pie.y = 25
    pie.width = 135
    pie.height = 135
    pie.data = [number(r.get("total", 0)) for r in rows]
    pie.labels = [""] * len(rows)
    pie.sideLabels = False
    pie.simpleLabels = False

    palette = [
        BLUE,
        GREEN,
        PURPLE,
        AMBER,
        CYAN,
        RED,
        HexColor("#EA580C"),
        HexColor("#7C3AED"),
    ]

    for i in range(len(rows)):
        pie.slices[i].fillColor = palette[i % len(palette)]
        pie.slices[i].strokeColor = colors.white
        pie.slices[i].strokeWidth = 1

    drawing.add(pie)

    drawing.add(
        String(
            100,
            90,
            "EXPENSES",
            textAnchor="middle",
            fontName=FONT_BOLD,
            fontSize=8,
            fillColor=MUTED,
        )
    )

    drawing.add(
        String(
            100,
            76,
            money(sum(number(r.get("total", 0)) for r in rows)),
            textAnchor="middle",
            fontName=FONT_BOLD,
            fontSize=10,
            fillColor=NAVY,
        )
    )

    # Legend
    y = 145
    for i, row in enumerate(rows[:7]):
        category = safe_text(row.get("category"), "Other")
        amount = money(row.get("total", 0))

        drawing.add(
            String(
                220,
                y,
                f"●  {category}",
                fontName=FONT_BOLD,
                fontSize=8,
                fillColor=palette[i % len(palette)],
            )
        )
        drawing.add(
            String(
                390,
                y,
                amount,
                fontName=FONT_REGULAR,
                fontSize=8,
                fillColor=NAVY,
            )
        )
        y -= 18

    drawing.add(
        String(
            300,
            165,
            "Spending by Category",
            textAnchor="middle",
            fontName=FONT_BOLD,
            fontSize=10,
            fillColor=NAVY,
        )
    )

    return drawing


def budget_progress(used, total):
    drawing = Drawing(500, 70)

    ratio = 0 if total <= 0 else min(max(used / total, 0), 1)

    drawing.add(
        String(
            0,
            55,
            "Budget used",
            fontName=FONT_BOLD,
            fontSize=8,
            fillColor=NAVY,
        )
    )

    drawing.add(
        String(
            500,
            55,
            f"{ratio * 100:.1f}%",
            textAnchor="end",
            fontName=FONT_BOLD,
            fontSize=8,
            fillColor=NAVY,
        )
    )

    drawing.add(
        Rect(
            0,
            28,
            500,
            14,
            fillColor=HexColor("#E2E8F0"),
            strokeColor=HexColor("#E2E8F0"),
            rx=7,
            ry=7,
        )
    )

    fill_color = RED if ratio > 1 else GREEN

    drawing.add(
        Rect(
            0,
            28,
            500 * ratio,
            14,
            fillColor=fill_color,
            strokeColor=fill_color,
            rx=7,
            ry=7,
        )
    )

    drawing.add(
        String(
            0,
            10,
            f"Used: {money(used)}",
            fontSize=7,
            fillColor=MUTED,
        )
    )

    drawing.add(
        String(
            500,
            10,
            f"Budget: {money(total)}",
            textAnchor="end",
            fontSize=7,
            fillColor=MUTED,
        )
    )

    return drawing


# ============================================================
# PAGE FOOTER
# ============================================================

def draw_footer(canvas, doc):
    canvas.saveState()

    width, height = A4

    canvas.setFillColor(NAVY)
    canvas.rect(
        0,
        height - 3 * mm,
        width,
        3 * mm,
        fill=1,
        stroke=0,
    )

    canvas.setStrokeColor(BORDER)
    canvas.setLineWidth(0.4)
    canvas.line(
        15 * mm,
        10 * mm,
        width - 15 * mm,
        10 * mm,
    )

    canvas.setFont(FONT_REGULAR, 7)
    canvas.setFillColor(MUTED)

    canvas.drawString(
        15 * mm,
        6 * mm,
        "BudgetBuddy • Personal Finance Report",
    )

    canvas.drawRightString(
        width - 15 * mm,
        6 * mm,
        f"Page {doc.page}",
    )

    canvas.restoreState()


# ============================================================
# JSON EXPORT
# ============================================================

class ExportJSONView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date, end_date = get_date_range(request)

        if start_date and end_date:
            if isinstance(start_date, str):
                start_date = datetime.strptime(
                    start_date,
                    "%Y-%m-%d",
                ).date()

            if isinstance(end_date, str):
                end_date = datetime.strptime(
                    end_date,
                    "%Y-%m-%d",
                ).date()

        if not start_date or not end_date:
            return HttpResponse(
                "Invalid Date Range",
                status=400,
            )

        report = build_report(
            request,
            start_date,
            end_date,
        )

        response = HttpResponse(
            json.dumps(
                report,
                indent=4,
                default=str,
            ),
            content_type="application/json",
        )

        response["Content-Disposition"] = (
            f'attachment; filename='
            f'"BudgetBuddy_Report_{start_date}_to_{end_date}.json"'
        )

        return response


# ============================================================
# PROFESSIONAL PDF EXPORT
# ============================================================

class ExportPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date, end_date = get_date_range(request)

        if start_date and end_date:
            if isinstance(start_date, str):
                start_date = datetime.strptime(
                    start_date,
                    "%Y-%m-%d",
                ).date()

            if isinstance(end_date, str):
                end_date = datetime.strptime(
                    end_date,
                    "%Y-%m-%d",
                ).date()

        if not start_date or not end_date:
            return HttpResponse(
                "Invalid Date Range",
                status=400,
            )

        if end_date < start_date:
            return HttpResponse(
                "End date must be greater than or equal to start date.",
                status=400,
            )

        report = build_report(
            request,
            start_date,
            end_date,
        )

        s = styles()
        story = []

        summary = report.get("summary", {})
        budget = report.get("budget", {})
        savings = report.get("savings", {})
        analytics = report.get("analytics", {})
        insights = report.get("insights", {})
        charts = report.get("charts", {})

        income = number(summary.get("total_income", 0))
        expense = number(summary.get("total_expense", 0))
        balance = number(summary.get("current_balance", 0))
        total_savings = number(summary.get("total_savings", 0))

        total_budget = number(budget.get("total_budget", 0))
        remaining_budget = number(budget.get("remaining_budget", 0))
        budget_used = number(
            analytics.get("budget_utilization", 0)
        )

        # --------------------------------------------------------
        # HEADER
        # --------------------------------------------------------

        # --------------------------------------------------------
        # BRAND HEADER
        # --------------------------------------------------------
        # Use the same final BudgetBuddy mark shown by the web app.
        # If the image is unavailable, the report still renders normally.
        logo_path = get_budgetbuddy_logo_path()

        if logo_path:
            logo = Image(
                str(logo_path),
                width=10 * mm,
                height=10 * mm,
            )

            brand = Table(
                [[
                    logo,
                    Paragraph("BudgetBuddy", s["title"]),
                ]],
                colWidths=[13 * mm, 112 * mm],
            )

            brand.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
        else:
            brand = Paragraph("BudgetBuddy", s["title"])

        header = Table(
            [[
                brand,
                Paragraph(
                    "<b>FINANCIAL REPORT</b><br/>"
                    "Personal finance overview",
                    ParagraphStyle(
                        "HeaderRight",
                        parent=s["subtitle"],
                        alignment=TA_RIGHT,
                        textColor=HexColor("#86EFAC"),
                    ),
                ),
            ]],
            colWidths=[125 * mm, 49 * mm],
        )

        header.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAVY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (0, 0), 11),
            ("RIGHTPADDING", (-1, 0), (-1, 0), 11),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))

        story.append(header)
        story.append(Spacer(1, 7))

        # Report metadata
        metadata = Table(
            [[
                Paragraph(
                    f"<b>REPORT PERIOD</b><br/>"
                    f"{start_date.strftime('%d %b %Y')} → "
                    f"{end_date.strftime('%d %b %Y')}",
                    s["body"],
                ),
                Paragraph(
                    f"<b>GENERATED</b><br/>"
                    f"{datetime.now().strftime('%d %b %Y, %I:%M %p')}",
                    s["body"],
                ),
                Paragraph(
                    "<b>ACCOUNT</b><br/>"
                    "Personal Finance",
                    s["body"],
                ),
            ]],
            colWidths=[66 * mm, 61 * mm, 47 * mm],
        )

        metadata.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BLUE),
            ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, BORDER),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        story.append(metadata)
        story.append(Spacer(1, 11))

        # --------------------------------------------------------
        # KEY NUMBERS
        # --------------------------------------------------------

        story.append(section_header("Financial snapshot", GREEN))
        story.append(Spacer(1, 5))

        snapshot = Table(
            [[
                stat_card("Income", money(income), GREEN),
                stat_card("Expenses", money(expense), RED),
                stat_card("Balance", money(balance), BLUE),
                stat_card("Savings", money(total_savings), PURPLE),
            ]],
            colWidths=[43.5 * mm] * 4,
        )

        snapshot.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))

        story.append(snapshot)
        story.append(Spacer(1, 10))

        # --------------------------------------------------------
        # INCOME VS EXPENSE
        # --------------------------------------------------------

        story.append(section_header("Income vs expenses", BLUE))
        story.append(Spacer(1, 2))
        story.append(income_expense_chart(income, expense))
        story.append(Spacer(1, 5))

        # One-line interpretation
        if income > expense:
            difference = income - expense
            comparison_text = (
                f"You earned <b>{money(difference)}</b> more than you spent "
                f"during this report period."
            )
            comparison_bg = HexColor("#F0FDF4")
        elif expense > income:
            difference = expense - income
            comparison_text = (
                f"You spent <b>{money(difference)}</b> more than you earned "
                f"during this report period."
            )
            comparison_bg = HexColor("#FEF2F2")
        else:
            comparison_text = "Income and expenses were equal."
            comparison_bg = LIGHT

        comparison = Table(
            [[Paragraph(comparison_text, s["body"])]],
            colWidths=[174 * mm],
        )

        comparison.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), comparison_bg),
            ("BOX", (0, 0), (-1, -1), 0.4, BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        story.append(comparison)
        story.append(Spacer(1, 10))

        # --------------------------------------------------------
        # BUDGET HEALTH
        # --------------------------------------------------------

        story.append(section_header("Budget health", GREEN))
        story.append(Spacer(1, 4))

        budget_summary = Table(
            [[
                Paragraph("<b>Total budget</b>", s["muted"]),
                Paragraph("<b>Remaining</b>", s["muted"]),
                Paragraph("<b>Budget used</b>", s["muted"]),
                Paragraph("<b>Status</b>", s["muted"]),
            ], [
                Paragraph(money(total_budget), s["card_value"]),
                Paragraph(money(remaining_budget), s["card_value"]),
                Paragraph(percent(budget_used), s["card_value"]),
                Paragraph(
                    f"<b>{safe_text(insights.get('budget_status'), 'Not available')}</b>",
                    s["body"],
                ),
            ]],
            colWidths=[43.5 * mm] * 4,
        )

        budget_summary.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        story.append(budget_summary)
        story.append(Spacer(1, 4))
        story.append(
            budget_progress(
                total_budget * budget_used / 100,
                total_budget,
            )
        )
        story.append(Spacer(1, 8))

        # --------------------------------------------------------
        # --------------------------------------------------------
        # PAGE 2
        # Keep the visual spending breakdown together with the
        # financial-health section so the report stays exactly 2 pages.
        # --------------------------------------------------------

        story.append(PageBreak())

        # --------------------------------------------------------
        # EXPENSE CATEGORY DONUT
        # --------------------------------------------------------

        story.append(section_header("Where you spent", PURPLE))
        story.append(Spacer(1, 2))

        story.append(
            expense_donut(
                charts.get("expense_by_category", [])
            )
        )
        story.append(Spacer(1, 3))

        story.append(section_header("Financial health", BLUE))
        story.append(Spacer(1, 4))

        health_cards = Table(
            [[
                stat_card(
                    "Transactions",
                    str(analytics.get("expense_transactions", 0)),
                    BLUE,
                ),
                stat_card(
                    "Savings rate",
                    percent(analytics.get("savings_rate", 0)),
                    GREEN,
                ),
                stat_card(
                    "Budget used",
                    percent(analytics.get("budget_utilization", 0)),
                    AMBER,
                ),
                stat_card(
                    "Goal completion",
                    percent(analytics.get("goal_completion_rate", 0)),
                    PURPLE,
                ),
            ]],
            colWidths=[43.5 * mm] * 4,
        )

        health_cards.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))

        story.append(health_cards)
        story.append(Spacer(1, 12))

        # --------------------------------------------------------
        # SAVINGS
        # --------------------------------------------------------

        story.append(section_header("Savings progress", GREEN))
        story.append(Spacer(1, 6))

        savings_table = Table(
            [[
                Paragraph("<b>Active goals</b>", s["muted"]),
                Paragraph("<b>Completed goals</b>", s["muted"]),
                Paragraph("<b>Savings rate</b>", s["muted"]),
                Paragraph("<b>Status</b>", s["muted"]),
            ], [
                Paragraph(
                    str(savings.get("active_goals", 0)),
                    s["card_value"],
                ),
                Paragraph(
                    str(savings.get("completed_goals", 0)),
                    s["card_value"],
                ),
                Paragraph(
                    percent(analytics.get("savings_rate", 0)),
                    s["card_value"],
                ),
                Paragraph(
                    f"<b>{safe_text(insights.get('savings_status'))}</b>",
                    s["body"],
                ),
            ]],
            colWidths=[43.5 * mm] * 4,
        )

        savings_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        story.append(savings_table)
        story.append(Spacer(1, 12))

        # --------------------------------------------------------
        # SIMPLE INSIGHTS
        # --------------------------------------------------------

        story.append(section_header("What this means", AMBER))
        story.append(Spacer(1, 5))

        highest = safe_text(
            insights.get("highest_expense_category"),
            "No expense category",
        )
        budget_status = safe_text(
            insights.get("budget_status"),
            "Not available",
        )
        savings_status = safe_text(
            insights.get("savings_status"),
            "Not available",
        )

        insight_text = (
            f"Your largest expense category was "
            f"<b>{highest}</b>. "
            f"Your budget is currently <b>{budget_status}</b>, "
            f"while your savings status is "
            f"<b>{savings_status}</b>."
        )

        insight_box = Table(
            [[Paragraph(insight_text, s["body"])]],
            colWidths=[174 * mm],
        )

        insight_box.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), HexColor("#FFFBEB")),
            ("BOX", (0, 0), (-1, -1), 0.5, HexColor("#FDE68A")),
            ("LINEBEFORE", (0, 0), (0, 0), 3, AMBER),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ]))

        story.append(insight_box)
        story.append(Spacer(1, 12))

        # --------------------------------------------------------
        # RECOMMENDATIONS
        # --------------------------------------------------------

        story.append(section_header("Recommended next steps", GREEN))
        story.append(Spacer(1, 5))

        recommendations = []

        if budget_status.lower() == "over budget":
            recommendations.append(
                "Review your largest expense categories and reduce avoidable spending."
            )
        else:
            recommendations.append(
                "Keep your current spending within the planned budget."
            )

        savings_rate = number(
            analytics.get("savings_rate", 0)
        )

        if savings_rate < 20:
            recommendations.append(
                "Try to increase your regular savings contribution."
            )
        else:
            recommendations.append(
                "Maintain your current savings habit."
            )

        if number(savings.get("completed_goals", 0)) == 0:
            recommendations.append(
                "Work toward completing your active savings goal."
            )
        else:
            recommendations.append(
                "Continue progressing toward your remaining savings goals."
            )

        rec_rows = [
            [Paragraph(f"<b>✓</b>  {item}", s["body"])]
            for item in recommendations
        ]

        recommendations_table = Table(
            rec_rows,
            colWidths=[174 * mm],
        )

        recommendations_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
            ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
            ("LINEBELOW", (0, 0), (-1, -2), 0.3, BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))

        story.append(recommendations_table)
        story.append(Spacer(1, 10))

        # Closing banner
        closing = Table(
            [[
                Paragraph(
                    "<b>BudgetBuddy</b><br/>"
                    "Simple numbers. Better financial decisions.",
                    s["body"],
                ),
                Paragraph(
                    "<b>PERSONAL</b><br/>"
                    "Financial report",
                    ParagraphStyle(
                        "ClosingRight",
                        parent=s["muted"],
                        alignment=TA_RIGHT,
                    ),
                ),
            ]],
            colWidths=[125 * mm, 49 * mm],
        )

        closing.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))

        story.append(closing)

        # --------------------------------------------------------
        # BUILD
        # --------------------------------------------------------

        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=14 * mm,
            bottomMargin=16 * mm,
            title="BudgetBuddy Financial Report",
            author="BudgetBuddy",
        )

        doc.build(
            story,
            onFirstPage=draw_footer,
            onLaterPages=draw_footer,
        )

        buffer.seek(0)

        return FileResponse(
            buffer,
            as_attachment=True,
            filename=(
                f"BudgetBuddy_Report_"
                f"{start_date}_to_{end_date}.pdf"
            ),
            content_type="application/pdf",
        )
# ============================================================
# EXCEL EXPORT
# ============================================================

class ExportExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date, end_date = get_date_range(request)

        if not start_date or not end_date:
            return HttpResponse(
                "start_date and end_date are required.",
                status=400,
                content_type="text/plain",
            )

        if end_date < start_date:
            return HttpResponse(
                "End date must be greater than or equal to start date.",
                status=400,
                content_type="text/plain",
            )

        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as ExcelImage
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse(
                "Excel export dependency is not installed. Run: pip install openpyxl",
                status=500,
                content_type="text/plain",
            )

        report = build_report(request, start_date, end_date)
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        navy = "0F172A"
        green = "16A34A"
        gold = "B8955A"
        light = "F8FAFC"
        border = "CBD5E1"

        thin = Side(style="thin", color=border)
        box = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Brand header
        logo_path = get_budgetbuddy_logo_path()
        if logo_path:
            try:
                logo = ExcelImage(str(logo_path))
                logo.width = 52
                logo.height = 52
                ws.add_image(logo, "A1")
            except Exception:
                pass

        ws.merge_cells("B1:F1")
        ws["B1"] = "BudgetBuddy"
        ws["B1"].font = Font(size=20, bold=True, color=navy)
        ws["B1"].alignment = Alignment(vertical="center")
        ws.merge_cells("B2:F2")
        ws["B2"] = "Personal Finance Report"
        ws["B2"].font = Font(size=11, color="64748B")
        ws.merge_cells("B3:F3")
        ws["B3"] = f"{start_date:%d %b %Y}  —  {end_date:%d %b %Y}"
        ws["B3"].font = Font(size=10, color="64748B")
        ws.row_dimensions[1].height = 34
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 22

        row = 5
        ws[f"A{row}"] = "Financial Summary"
        ws[f"A{row}"].font = Font(size=14, bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=navy)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        summary_rows = [
            ("Total Income", report["summary"]["total_income"]),
            ("Total Expenses", report["summary"]["total_expense"]),
            ("Current Balance", report["summary"]["current_balance"]),
            ("Total Savings", report["summary"]["total_savings"]),
            ("Total Budget", report["budget"]["total_budget"]),
            ("Budget Remaining", report["budget"]["remaining_budget"]),
            ("Overspent", report["budget"]["overspent_amount"]),
            ("Savings Rate", f'{report["analytics"]["savings_rate"]}%'),
            ("Budget Utilization", f'{report["analytics"]["budget_utilization"]}%'),
        ]
        for label, value in summary_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=float(value) if hasattr(value, "as_tuple") else value)
            ws.cell(row=row, column=1).border = box
            ws.cell(row=row, column=2).border = box
            row += 1

        # Expense categories
        row += 1
        ws.cell(row=row, column=1, value="Expense by Category")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=green)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for col, value in enumerate(("Category", "Amount", "Share"), 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.border = box
        row += 1
        total_expense = float(report["summary"]["total_expense"] or 0)
        for item in report["charts"]["expense_by_category"]:
            amount = float(item.get("total") or 0)
            values = [item.get("category", "Other"), amount, f"{(amount / total_expense * 100):.1f}%" if total_expense else "0.0%"]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = box
            row += 1

        # Transactions sheets
        income_ws = wb.create_sheet("Income")
        income_ws.append(["Date", "Source", "Amount", "Description"])
        for cell in income_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
        for item in report["transactions"]["income"]:
            income_ws.append([item["income_date"], item["source"], float(item["amount"] or 0), item.get("description", "")])

        expense_ws = wb.create_sheet("Expenses")
        expense_ws.append(["Date", "Category", "Amount", "Description"])
        for cell in expense_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
        for item in report["transactions"]["expenses"]:
            expense_ws.append([item["expense_date"], item["category"], float(item["amount"] or 0), item.get("description", "")])

        goals_ws = wb.create_sheet("Savings Goals")
        goals_ws.append(["Goal", "Target", "Saved", "Remaining", "Progress", "Status"])
        for cell in goals_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
        for goal in report["savings"]["goals"]:
            goals_ws.append([
                goal.get("goal_name"),
                float(goal.get("target_amount") or 0),
                float(goal.get("saved_amount") or 0),
                float(goal.get("remaining_amount") or 0),
                f'{goal.get("progress_percentage", 0)}%',
                goal.get("status"),
            ])

        for sheet in wb.worksheets:
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 38)
            sheet.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="BudgetBuddy_Report_{start_date}_to_{end_date}.xlsx"'
        )
        return response
